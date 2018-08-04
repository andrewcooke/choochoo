
import datetime as dt
from abc import abstractmethod
from os.path import dirname, join
from pickle import dump, load
from re import compile
from struct import unpack

import openpyxl as xls
from more_itertools import peekable
from pkg_resources import resource_stream

from .records import LazyRecord
from ..args import PATH
from ..log import make_log

LITTLE, BIG = 0, 1
PROFILE = 'global-profile.pkl'
HEADER_GLOBAL_TYPE = -1
TIMESTAMP_GLOBAL_TYPE = 253
HEADER_FIELDS = [
    ('header_size', 1, 'uint8'),
    ('protocol_version', 1, 'uint8'),
    ('profile_version', 1, 'uint16'),
    ('data_size', 1, 'uint32'),
    ('fit_text', 4, 'string'),
    ('checksum', 1, 'uint16')
]


def package_fit_profile(args):
    log = make_log(args)
    in_path = args.file(PATH, 0)
    log.info('Reading from %s' % in_path)
    nlog, types, messages = read_profile(log, in_path)
    out_path = join(dirname(__file__), PROFILE)
    nlog.set_log(None)
    log.info('Writing to %s' % out_path)
    with open(out_path, 'wb') as output:
        dump((nlog, types, messages), output)
    # test loading
    log.info('Test loading from %r' % PROFILE)
    log.info('Loaded %s, %s' % load_profile(log))


def read_profile(log, path):
    nlog = NullableLog(log)
    wb = xls.load_workbook(path)
    types = Types(nlog, wb['Types'])
    messages = Messages(nlog, wb['Messages'], types)
    return nlog, types, messages


def load_profile(log):
    input = resource_stream(__name__, PROFILE)
    nlog, types, messages = load(input)
    nlog.set_log(log)
    return types, messages


class NullableLog:

    def __init__(self, log):
        self.set_log(log)

    def set_log(self, log):
        self.__log = log

    def debug(self, *args):
        self.__log.debug(*args)

    def info(self, *args):
        self.__log.info(*args)

    def warn(self, *args):
        self.__log.warn(*args)

    def error(self, *args):
        self.__log.error(*args)


class Named:
    """
    Has a name.  Base for both fields and messages
    """

    def __init__(self, log, name):
        self._log = log
        self.name = name

    def __str__(self):
        return '%s: %s' % (self.__class__.__name__, self.name)


class ErrorDict(dict):

    def __init__(self, log, error_msg):
        self.__log = log
        self.__error_msg = error_msg
        super().__init__()

    def add_named(self, item):
        self[item.name] = item

    def __getitem__(self, item):
        try:
            return super().__getitem__(item)
        except KeyError:
            msg = self.__error_msg % (item,)
            self.__log.error(msg)
            raise KeyError(msg)


class ErrorList(list):

    def __init__(self, log, error_msg):
        self.__log = log
        self.__error_msg = error_msg
        super().__init__()

    def __getitem__(self, item):
        try:
            return super().__getitem__(item)
        except IndexError:
            msg = self.__error_msg % item
            self.__log.error(msg)
            raise IndexError(msg)


class AbstractType(Named):

    def __init__(self, log, name, size, base_type=None):
        super().__init__(log, name)
        self.base_type = base_type
        self.size = size

    @abstractmethod
    def profile_to_internal(self, cell_contents):
        raise NotImplementedError('%s: %s' % (self.__class__.__name__, self.name))

    @abstractmethod
    def parse(self, bytes, count, endian):
        raise NotImplementedError('%s: %s' % (self.__class__.__name__, self.name))


class BaseType(AbstractType):

    def __init__(self, log, name, size, func):
        super().__init__(log, name, size)
        self.__func = func

    def profile_to_internal(self, cell_contents):
        return self.__func(cell_contents)


class StructSupport(BaseType):

    def _pack_bad(self, value):
        bad = (bytearray(self.size), bytearray(self.size))
        for endian in (LITTLE, BIG):
            bytes = value
            for i in range(self.size):
                j = i if endian == LITTLE else self.size - i - 1
                bad[endian][j] = bytes & 0xff
                bytes >>= 8
        return bad

    def _is_bad(self, data, bad):
        size = len(bad)
        count = len(data) // size
        return all(bad == data[size*i:size*(i+1)] for i in range(count))

    def _unpack(self, data, formats, bad, count, endian):
        if self._is_bad(data, bad[endian]):
            return None
        else:
            return unpack(formats[endian] % count, data[0:count * self.size])


class String(BaseType):

    # it may seem weird this returns a singleton tuple.  shouldn't it be treated like a
    # Character field where count is the mutliplicity?  but dynamic fields treat it as
    # a single value...

    def __init__(self, log, name):
        super().__init__(log, name, 1, str)

    def parse(self, bytes, count, endian):
        return (str(b''.join(unpack('%dc' % count, bytes)), encoding='utf-8'),)


class Boolean(BaseType):

    def __init__(self, log, name):
        super().__init__(log, name, 1, bool)

    def parse(self, bytes, count, endian):
        return tuple(bool(byte) for byte in bytes)


class AutoInteger(StructSupport):

    pattern = compile(r'^([su]?)int(\d{1,2})(z?)$')

    size_to_format = {1: 'b', 2: 'h', 4: 'i', 8: 'q'}

    def __init__(self, log, name):
        match = self.pattern.match(name)
        self.signed = match.group(1) != 'u'
        bits = int(match.group(2))
        if bits % 8:
            raise Exception('Size of %r not a multiple of 8 bits' % name)
        super().__init__(log, name, bits // 8, self.int)
        if self.size not in self.size_to_format:
            raise Exception('Cannot unpack %d bytes as an integer' % self.size)
        format = self.size_to_format[self.size]
        if not self.signed:
            format = format.upper()
        self.formats = ['<%d' + format, '>%d' + format]
        self.bad = self._pack_bad(0 if match.group(3) == 'z' else 2 ** (bits - (1 if self.signed else 0)) - 1)

    @staticmethod
    def int(cell):
        if isinstance(cell, int):
            return cell
        else:
            return int(cell, 0)

    def parse(self, data, count, endian):
        result = self._unpack(data, self.formats, self.bad, count, endian)
        if result is not None and self.size == 1:
            result = bytes(result)
        return result


class AliasInteger(AutoInteger):

    def __init__(self, log, name, spec):
        super().__init__(log, spec)
        self.name = name


class Date(AliasInteger):

    def __init__(self, log, name, utc, to_datetime=True):
        super().__init__(log, name, 'uint32')
        self.__tzinfo = dt.timezone.utc if utc else None
        self.__to_datetime = to_datetime

    @staticmethod
    def convert(time, tzinfo=dt.timezone.utc):
        if time >= 0x10000000 :
            return dt.datetime(1989, 12, 31, tzinfo=tzinfo) + dt.timedelta(seconds=time)
        else:
            return time

    def parse(self, data, count, endian):
        times = super().parse(data, count, endian)
        if self.__to_datetime:
            times = tuple(self.convert(time, tzinfo=self.__tzinfo) for time in times)
        return times


class AutoFloat(StructSupport):

    pattern = compile(r'^float(\d{1,2})$')

    size_to_format = {2: 'e', 4: 'f', 8: 'd'}

    def __init__(self, log, name):
        match = self.pattern.match(name)
        bits = int(match.group(1))
        if bits % 8:
            raise Exception('Size of %r not a multiple of 8 bits' % name)
        super().__init__(log, name, bits // 8, float)
        if self.size not in self.size_to_format:
            raise Exception('Cannot unpack %d bytes as a float' % self.size)
        format = self.size_to_format[self.size]
        self.formats = ['<%d' + format, '>%d' + format]
        self.bad = self._pack_bad(2 ** bits - 1)

    def parse(self, data, count, endian):
        return self._unpack(data, self.formats, self.bad, count, endian)


class Mapping(AbstractType):

    def __init__(self, log, row, rows, types):
        name = row[0]
        base_type_name = row[1]
        base_type = types.profile_to_type(base_type_name, auto_create=True)
        super().__init__(log, name, base_type.size, base_type=base_type)
        self._profile_to_internal = ErrorDict(log, 'No internal value for profile %r')
        self._internal_to_profile = ErrorDict(log, 'No profile value for internal %r')
        for row in rows:
            if row[0] or row[2] is None or row[3] is None:
                rows.prepend(row)
                break
            self.__add_mapping(row)
        # log.debug('Parsed %d values' % len(self._profile_to_internal))

    def profile_to_internal(self, cell_contents):
        return self._profile_to_internal[cell_contents]

    def internal_to_profile(self, value):
        return self._internal_to_profile[value]

    def safe_internal_to_profile(self, value):
        try:
            return self.internal_to_profile(value)
        except KeyError:
            return value

    def parse(self, bytes, size, endian):
        values = self.base_type.parse(bytes, size, endian)
        if values:
            values = tuple(self.safe_internal_to_profile(value) for value in values)
        return values

    def __add_mapping(self, row):
        profile = row[2]
        internal = self.base_type.profile_to_internal(row[3])
        self._profile_to_internal[profile] = internal
        self._internal_to_profile[internal] = profile


# table 4-6 of FIT defn doc
BASE_TYPE_NAMES = ['enum', 'sint8', 'uint8', 'sint16', 'uint16', 'sint32', 'uint32',
                   'string', 'float32', 'float64',
                   'uint8z', 'uint16z', 'uint32z', 'byte', 'sint64', 'uint64', 'uint64z']


class Types:

    def __init__(self, log, sheet, to_datetime=True):
        self.__log = log
        self.__profile_to_type = ErrorDict(log, 'No type for profile %r')
        # these are not 'base types' in the same sense as types having base types.
        # rather, they are the 'base (integer) types' described in the docs
        self.base_types = ErrorList(log, 'No base type for number %r')
        self.__add_known_types(to_datetime)
        rows = peekable([cell.value for cell in row] for row in sheet.iter_rows())
        for row in rows:
            if row[0] and row[0][0].isupper():
                self.__log.debug('Skipping %s' % row)
            elif row[0]:
                # self.__log.info('Parsing type %s' % row[0])
                self.__add_type(Mapping(self.__log, row, rows, self))

    def __add_known_types(self, to_datetime):
        # these cannot be inferred from name
        self.__add_type(String(self.__log, 'string'))
        self.__add_type(AliasInteger(self.__log, 'enum', 'uint8'))
        self.__add_type(AliasInteger(self.__log, 'byte', 'uint8'))
        # these can be inferred
        for name in BASE_TYPE_NAMES:
            self.profile_to_type(name, auto_create=True)
            self.base_types.append(self.profile_to_type(name))
        # this is in the spreadsheet, but not in the doc
        self.__add_type(Boolean(self.__log, 'bool'))
        # these are defined in the spreadsheet, but the interpretation is in comments
        self.__add_type(Date(self.__log, 'date_time', True, to_datetime=to_datetime))
        self.__add_type(Date(self.__log, 'local_date_time', False, to_datetime=to_datetime))

    def __add_type(self, type):
        if type.name in self.__profile_to_type:
            duplicate = self.__profile_to_type[type.name]
            if duplicate.size == type.size:
                self.__log.warn('Ignoring duplicate type for %r' % type.name)
            else:
                raise Exception('Duplicate type for %r with differing size (%d  %d)' %
                                (type.name, type.size, duplicate.size))
        else:
            self.__profile_to_type.add_named(type)

    def profile_to_type(self, name, auto_create=False):
        try:
            return self.__profile_to_type[name]
        except KeyError:
            if auto_create:
                for cls in (AutoFloat, AutoInteger):
                    match = cls.pattern.match(name)
                    if match:
                        self.__log.warn('Auto-adding type %s for %r' % (cls.__name__, name))
                        self.__add_type(cls(self.__log, name))
                        return self.profile_to_type(name)
            raise


def scale_offset(log, cell, default, name):
    if cell is None or cell == '':
        return default
    try:
        return int(cell)
    except:
        log.warn('Could not parse %r for %s (scale/offset)' % (cell, name))
        return default


class MessageField(Named):

    def __init__(self, log, name, number, units, type, count=None, scale=1, offset=0):
        super().__init__(log, name)
        self.number = number
        self.units = units if units else ''
        self.is_dynamic = False
        self.type = type
        self.count = count
        self.scale = scale_offset(log, scale, 1, name)
        self.offset = scale_offset(log, offset, 0, name)
        self.__is_scaled = (self.scale != 1 or self.offset != 0)

    def parse(self, data, count, endian, result, message):
        values = self.type.parse(data, count, endian)
        if self.__is_scaled and values is not None:
            values = tuple(value / self.scale - self.offset for value in values)
        return self.name, (values, self.units)


class RowMessageField(MessageField):

    def __init__(self, log, row, types):
        super().__init__(log, row[2],
                         int(row[1]) if row[1] is not None else None,
                         row[8],
                         types.profile_to_type(row[3], auto_create=True),
                         row[15], row[6], row[7])


class DynamicMessageField(RowMessageField):

    # actually a possible dynamic field - see is_dynamic

    def __init__(self, log, row, rows, types):
        super().__init__(log, row, types)
        self.__dynamic_tmp_data = []
        self.__dynamic_lookup = ErrorDict(log, 'No dynamic field for %r')
        self.references = set()
        try:
            peek = rows.peek()
            while peek[2] and peek[1] is None:
                row = next(rows)
                for name, value in zip(row[11].split(','), row[12].split(',')):
                    self.__save_dynamic(name.strip(), value.strip(), row)
                peek = rows.peek()
        except StopIteration:
            return

    def __save_dynamic(self, reference_name, reference_value, row):
        self.is_dynamic = True
        self.__dynamic_tmp_data.append((reference_name, reference_value, row))

    def _complete_dynamic(self, message, types):
        for name, value, row in self.__dynamic_tmp_data:
            self.references.add(name)
            self.__dynamic_lookup[(name, value)] = RowMessageField(self._log, row, types)

    @property
    def dynamic(self):
        return self.__dynamic_lookup

    def parse(self, data, count, endian, result, message):
        if self.is_dynamic:
            for name in self.references:
                if name in result:
                    value = result[name][0][0]  # drop units and take first value
                    self._log.debug('Found reference %r=%r' % (name, value))
                    try:
                        return self.dynamic[(name, value)].parse(data, count, endian, result, message)
                    except KeyError:
                        pass
            # self._log.warn('No match for dynamic field %s (message %s)' % (self.name, message.name))
            for option, field in self.__dynamic_lookup.items():
                # self._log.debug('Option: %s -> %r' % (option, field.name))
                pass
            # and if nothing found, fall though to default behaviour
        return super().parse(data, count, endian, result, message)


class Message(Named):

    def __init__(self, log, name, number=None):
        super().__init__(log, name)
        self.number = number
        self._profile_to_field = ErrorDict(log, 'No field for profile %r')
        self._number_to_field = ErrorDict(log, 'No field for number %r')

    def _add_field(self, field):
        self._profile_to_field.add_named(field)
        self._number_to_field[field.number] = field

    def profile_to_field(self, name):
        return self._profile_to_field[name]

    def number_to_field(self, value):
        return self._number_to_field[value]

    def parse(self, data, defn, timestamp=None):
        return LazyRecord(self.name, self.number, defn.identity, timestamp, self.__parse(data, defn))

    def __parse(self, data, defn):
        references = {} if defn.references else None
        # note this is a field description not a message field
        for field in defn.fields:
            bytes = data[field.start:field.finish]
            if field.field:
                name, value = self._parse_field(
                    field.field, bytes, field.count, defn.endian, references, self)
            else:
                name = str(field.number)
                value = (field.base_type.parse(bytes, field.count, defn.endian), None)
            if name in defn.references:
                references[name] = value
            yield name, value

    def _parse_field(self, field, bytes, count, endian, references, message):
        # allow interception for optional field in header
        return field.parse(bytes, count, endian, references, message)


class NumberedMessage(Message):

     def __init__(self, log, name, types):
        try:
            number = types.profile_to_type('mesg_num').profile_to_internal(name)
        except KeyError:
            number = None
            log.warn('No mesg_num for %r' % name)
        super().__init__(log, name, number)


class RowMessage(NumberedMessage):

    def __init__(self, log, row, rows, types):
        super().__init__(log, row[0], types)
        for row in rows:
            if not row[2]:
                rows.prepend(row)
                break
            self.__parse_row(row, rows, types)
        self.__complete_dynamic(types)

    def __parse_row(self, row, rows, types):
        self._add_field(DynamicMessageField(self._log, row, rows, types))

    def __complete_dynamic(self, types):
        # these may be forward references
        for field in self._profile_to_field.values():
            if field.is_dynamic:
                field._complete_dynamic(self, types)


class Header(Message):

    def __init__(self, log, types):
        super().__init__(log, 'HEADER', number=HEADER_GLOBAL_TYPE)
        for n, (name, size, base_type) in enumerate(HEADER_FIELDS):
            self._add_field(MessageField(log, name, n, None, types.profile_to_type(base_type)))

    def _parse_field(self, field, data, count, endian, references, message):
        if field.name == 'checksum' and references['header_size'] == 12:
            return None, None
        else:
            return super()._parse_field(field, data, count, endian, references, message)


class Missing(Message):

    def __init__(self, log, number):
        super().__init__(log, 'MESSAGE %d' % number, number)


class Messages:

    def __init__(self, log, sheet, types):
        self.__log = log
        self.__profile_to_message = ErrorDict(log, 'No message for profile %r')
        self.__number_to_message = ErrorDict(log, 'No message for number %r')
        rows = peekable([cell.value for cell in row] for row in sheet.iter_rows())
        for row in rows:
            if row[0] and row[0][0].isupper():
                self.__log.debug('Skipping %s' % row)
            elif row[0]:
                # self.__log.info('Parsing message %s' % row[0])
                self.__add_message(RowMessage(self.__log, row, rows, types))
        self.__add_message(Header(self.__log, types))

    def __add_message(self, message):
        self.__profile_to_message.add_named(message)
        try:
            self.__number_to_message[message.number] = message
        except AttributeError:
            pass

    def profile_to_message(self, name):
        return self.__profile_to_message[name]

    def number_to_message(self, number):
        try:
            return self.__number_to_message[number]
        except KeyError:
            message = Missing(self.__log, number)
            self.__number_to_message[number] = message
            return message
